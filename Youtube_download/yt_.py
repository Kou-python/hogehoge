from yt_dlp import YoutubeDL
import openpyxl
import time
from time import sleep

directry='ディレクトリを指定して下さい。'

# ブック取得
wb = openpyxl.load_workbook(directry+'保存リスト.xlsx', data_only=True)

# シートを取得
sheet = wb["Sheet1"]
max_row=sheet.max_row+1

# ExcelからURLを読み取り、動画を連続でダウンロードする
for i in range(1,max_row):
    URL=(sheet.cell(row=i, column=1).value) #URLを読み込み
    print(URL)
    
    ydl_opts = {
                'format': 'best',
                'outtmpl': '%(title)s.%(ext)s'
                }
    with YoutubeDL(ydl_opts) as ydl:
        result = ydl.download([URL])
        print(result)
    sleep(2)